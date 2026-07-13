"""Parse flood water time-series Excel into JSON for the frontend."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\Kunal.Desale\Downloads\630ca7ab6a874dfdaad92c98139edb07 (1).xlsx")
OUT = Path(r"C:\Users\Kunal.Desale\Desktop\Adani\frontend\public\flood_timeseries.json")


def _f(v):
    try:
        if v is None:
            return None
        return float(v)
    except (TypeError, ValueError):
        return None


def _s(v):
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()[:10]


def main() -> None:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

    # Aggregate series (one row per observation date; prefer sheet-matched metrics)
    ts_ws = wb["timeseries"]
    ts_rows = list(ts_ws.iter_rows(values_only=True))
    ts_headers = [str(h).strip() if h else "" for h in ts_rows[0]]
    ti = {h: i for i, h in enumerate(ts_headers)}

    by_date: dict[str, dict] = {}
    for row in ts_rows[1:]:
        if not row:
            continue
        date = _s(row[ti["post_date"]])
        if not date:
            continue
        entry = {
            "date": date,
            "pre_date": _s(row[ti["pre_date"]]),
            "lat": _f(row[ti["latitude"]]),
            "lon": _f(row[ti["longitude"]]),
            "water_area_ha": _f(row[ti["water_area_ha"]]),
            "flood_area_ha": _f(row[ti["flood_area_ha"]]),
        }
        # Keep last occurrence (sheet 2026-06-25 matches the later row)
        by_date[date] = entry

    scenes: dict[str, dict] = {}
    for name in wb.sheetnames:
        if name == "timeseries":
            continue
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else "" for h in rows[0]]
        hi = {h: i for i, h in enumerate(headers)}
        points = []
        water_n = flood_n = 0
        for row in rows[1:]:
            if not row:
                continue
            lat = _f(row[hi.get("latitude", -1)]) if "latitude" in hi else None
            lon = _f(row[hi.get("longitude", -1)]) if "longitude" in hi else None
            if lat is None or lon is None:
                continue
            cls = str(row[hi["class"]]).strip().lower() if "class" in hi and row[hi["class"]] else "water"
            if cls not in ("water", "flood"):
                cls = "water"
            if cls == "flood":
                flood_n += 1
            else:
                water_n += 1
            points.append({"lat": round(lat, 6), "lon": round(lon, 6), "class": cls})

        date = name if name in by_date else _s(name)
        meta = by_date.get(date, {"date": date})
        scenes[date] = {
            "date": date,
            "pre_date": meta.get("pre_date"),
            "water_area_ha": meta.get("water_area_ha"),
            "flood_area_ha": meta.get("flood_area_ha"),
            "water_points": water_n,
            "flood_points": flood_n,
            "points": points,
        }

    dates = sorted(scenes.keys())
    timeseries = []
    for d in dates:
        sc = scenes[d]
        base = by_date.get(d, {})
        timeseries.append(
            {
                "date": d,
                "pre_date": sc.get("pre_date") or base.get("pre_date"),
                "lat": base.get("lat"),
                "lon": base.get("lon"),
                "water_area_ha": sc.get("water_area_ha") if sc.get("water_area_ha") is not None else base.get("water_area_ha"),
                "flood_area_ha": sc.get("flood_area_ha") if sc.get("flood_area_ha") is not None else base.get("flood_area_ha"),
                "water_points": sc["water_points"],
                "flood_points": sc["flood_points"],
            }
        )

    payload = {
        "title": "Flood water time series",
        "description": "Satellite-derived water and flood extent along the corridor (ha)",
        "unit": "ha",
        "dates": dates,
        "timeseries": timeseries,
        "scenes": scenes,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Wrote {len(dates)} dates, {sum(len(s['points']) for s in scenes.values())} points -> {OUT}")
    wb.close()


if __name__ == "__main__":
    main()
