"""Parse tree inventory Excel into JSON for the frontend map overlay."""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl

SRC = Path(r"C:\Users\Kunal.Desale\Downloads\de7eb80b6eb74d42abad45330a1f0a7d.xlsx")
OUT = Path(r"C:\Users\Kunal.Desale\Desktop\Adani\frontend\public\trees.json")


def main() -> None:
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}

    def col(*names: str) -> int | None:
        for n in names:
            if n in idx:
                return idx[n]
        return None

    i_sr = col("Sr_No", "sr_no")
    i_id = col("pin_id", "id", "tree_id")
    i_lat = col("latitude", "lat")
    i_lon = col("longitude", "lon", "lng")
    i_zone = col("zone")
    i_ch = col("chainage_start_m", "chainage_m", "chainage")

    trees: list[dict] = []
    for row in rows:
        if not row:
            continue
        try:
            lat = float(row[i_lat]) if i_lat is not None and row[i_lat] is not None else None
            lon = float(row[i_lon]) if i_lon is not None and row[i_lon] is not None else None
        except (TypeError, ValueError):
            continue
        if lat is None or lon is None:
            continue
        if not (-90 <= lat <= 90 and -180 <= lon <= 180):
            continue

        pin = str(row[i_id]).strip() if i_id is not None and row[i_id] is not None else None
        zone = str(row[i_zone]).strip() if i_zone is not None and row[i_zone] is not None else None
        chainage_m = None
        if i_ch is not None and row[i_ch] is not None:
            try:
                chainage_m = float(row[i_ch])
            except (TypeError, ValueError):
                chainage_m = None

        sr = None
        if i_sr is not None and row[i_sr] is not None:
            try:
                sr = int(row[i_sr])
            except (TypeError, ValueError):
                sr = None

        trees.append(
            {
                "id": pin or f"Tree_{len(trees) + 1:04d}",
                "sr": sr,
                "lat": round(lat, 7),
                "lon": round(lon, 7),
                "zone": zone,
                "chainage_m": chainage_m,
            }
        )

    wb.close()
    payload = {
        "title": "Trees",
        "description": "Tree inventory along the project corridor",
        "count": len(trees),
        "trees": trees,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
    print(f"Wrote {len(trees)} trees -> {OUT}")


if __name__ == "__main__":
    main()
