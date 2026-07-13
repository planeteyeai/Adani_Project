"""
Parse Schedule-B engineering tables from the bundled Excel workbook.

The workbook covers crust composition, typical cross sections (TCS), paved
shoulders, underpasses, overpasses, service roads, elevated sections,
interchanges, drains, culverts and RE walls for the Digha–Koilwar corridor.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment,misc]

_NUM = re.compile(r"^-?\d+(?:\.\d+)?$")


def _num(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text or text.lower() in {"none", "—", "-"}:
        return None
    if _NUM.match(text):
        return float(text)
    return None


def _text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _row_has_serial(row: tuple, col: int = 0) -> bool:
    val = row[col] if col < len(row) else None
    if isinstance(val, (int, float)) and val > 0 and val == int(val):
        return True
    # Some sheets put S. No. in column 1 with a blank first cell.
    if col == 0 and val is None and len(row) > 1:
        alt = row[1]
        return isinstance(alt, (int, float)) and alt > 0 and alt == int(alt)
    return False


def _serial_col(row: tuple) -> int:
    val0 = row[0] if row else None
    if isinstance(val0, (int, float)) and val0 > 0 and val0 == int(val0):
        return 0
    val1 = row[1] if len(row) > 1 else None
    if isinstance(val1, (int, float)) and val1 > 0 and val1 == int(val1):
        return 1
    return 0


def _find_header_row(rows: list[tuple], marker: str) -> int | None:
    marker_l = marker.lower()
    for i, row in enumerate(rows):
        joined = " ".join(str(c or "") for c in row).lower()
        if marker_l in joined:
            return i
    return None


def parse_schedule_b(path: str | Path) -> dict:
    if load_workbook is None:
        raise ImportError("openpyxl is required to parse Schedule-B Excel files")

    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)

    def sheet_rows(name: str) -> list[tuple]:
        ws = wb[name]
        return [tuple(r) for r in ws.iter_rows(values_only=True)]

    data: dict[str, Any] = {
        "source_file": path.name,
        "crust_layers": _parse_crust(sheet_rows("Crust Layer")),
        "tcs_stretches": _parse_tcs(sheet_rows("TCS")),
        "paved_shoulders": _parse_paved_shoulder(sheet_rows("Paved Shoulder")),
        "underpasses": _parse_underpasses(sheet_rows("Underpasses")),
        "overpasses": _parse_overpasses(sheet_rows("Overpasses,Cloverleaf")),
        "service_roads": _parse_service_roads(sheet_rows("SR,Slip Road.Connecting Road")),
        "elevated": _parse_elevated(sheet_rows("Elevated")),
        "interchanges": _parse_interchanges(sheet_rows("Interchange (IC)")),
        "interchange_ramps": _parse_interchange_ramps(sheet_rows("Interchange (IC)")),
        "drains": _parse_drains(sheet_rows("Drain")),
        "culverts": _parse_culverts(sheet_rows("New Culvert")),
        "re_walls": _parse_re_walls(sheet_rows("RE Wall")),
    }

    data["structures"] = _build_structure_inventory(data)
    data["summary"] = _build_summary(data)
    wb.close()
    return data


def _parse_crust(rows: list[tuple]) -> dict:
    main: list[dict] = []
    service: list[dict] = []
    section = "main"
    for row in rows:
        cells = [c for c in row if c is not None]
        if not cells:
            continue
        joined = " ".join(str(c) for c in cells).lower()
        if "crossroads" in joined or "service roads" in joined:
            section = "service"
            continue
        if not _row_has_serial(row):
            continue
        layer = _text(row[2]) if len(row) > 2 else None
        thickness = _text(row[3]) if len(row) > 3 else None
        if not layer:
            continue
        entry = {"layer": layer, "thickness": thickness}
        (main if section == "main" else service).append(entry)
    return {"main_carriageway": main, "service_roads": service}


def _parse_tcs(rows: list[tuple]) -> list[dict]:
    stretches: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        from_km = _num(row[1]) if len(row) > 1 else None
        to_km = _num(row[2]) if len(row) > 2 else None
        if from_km is None or to_km is None:
            continue
        stretches.append(
            {
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "length_km": round(_num(row[3]) or (to_km - from_km), 3),
                "tcs": _text(row[4]) if len(row) > 4 else None,
                "description": _text(row[5]) if len(row) > 5 else None,
                "remarks": _text(row[6]) if len(row) > 6 else None,
            }
        )
    return stretches


def _parse_paved_shoulder(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        from_km = _num(row[1]) if len(row) > 1 else None
        to_km = _num(row[2]) if len(row) > 2 else None
        if from_km is None or to_km is None:
            continue
        items.append(
            {
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "length_km": round(_num(row[3]) or (to_km - from_km), 3),
                "shoulder_type": _text(row[4]) if len(row) > 4 else None,
                "tcs_ref": _text(row[5]) if len(row) > 5 else None,
            }
        )
    return items


def _parse_underpasses(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        chainage = _num(row[1]) if len(row) > 1 else None
        if chainage is None:
            continue
        items.append(
            {
                "chainage_km": round(chainage, 3),
                "lhs_width_m": _num(row[2]) if len(row) > 2 else None,
                "rhs_width_m": _num(row[3]) if len(row) > 3 else None,
                "median_superstructure": _text(row[4]) if len(row) > 4 else None,
                "span_arrangement": _text(row[5]) if len(row) > 5 else None,
                "vertical_clearance_m": _num(row[6]) if len(row) > 6 else None,
                "superstructure_type": _text(row[7]) if len(row) > 7 else None,
                "total_width_m": _text(row[8]) if len(row) > 8 else None,
                "remarks": _text(row[9]) if len(row) > 9 else None,
                "skew_angle": _text(row[10]) if len(row) > 10 else None,
            }
        )
    return items


def _parse_overpasses(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        chainage = _num(row[1]) if len(row) > 1 else None
        if chainage is None:
            continue
        items.append(
            {
                "chainage_km": round(chainage, 3),
                "lhs_width_m": _num(row[2]) if len(row) > 2 else None,
                "rhs_width_m": _num(row[3]) if len(row) > 3 else None,
                "median_superstructure": _text(row[4]) if len(row) > 4 else None,
                "span_arrangement": _text(row[5]) if len(row) > 5 else None,
                "vertical_clearance_m": _num(row[6]) if len(row) > 6 else None,
                "superstructure_type": _text(row[7]) if len(row) > 7 else None,
                "total_width_m": _text(row[8]) if len(row) > 8 else None,
                "remarks": _text(row[9]) if len(row) > 9 else None,
            }
        )
    return items


def _parse_service_roads(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        col = _serial_col(row)
        from_km = _num(row[col + 1]) if len(row) > col + 1 else None
        to_km = _num(row[col + 2]) if len(row) > col + 2 else None
        if from_km is None or to_km is None:
            continue
        items.append(
            {
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "lhs_length_km": _num(row[col + 3]) if len(row) > col + 3 else None,
                "rhs_length_km": _num(row[col + 4]) if len(row) > col + 4 else None,
                "carriageway_width": _text(row[col + 5]) if len(row) > col + 5 else None,
                "total_length_km": _num(row[col + 6]) if len(row) > col + 6 else None,
                "remarks": _text(row[col + 7]) if len(row) > col + 7 else None,
            }
        )
    return items


def _parse_elevated(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        from_km = _num(row[1]) if len(row) > 1 else None
        to_km = _num(row[2]) if len(row) > 2 else None
        if from_km is None or to_km is None:
            continue
        length_raw = _num(row[3]) if len(row) > 3 else None
        length_km = round(to_km - from_km, 3)
        if length_raw is not None and length_raw > 100:
            length_m = length_raw
        else:
            length_m = round(length_km * 1000, 1) if length_km else None
        items.append(
            {
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "length_km": length_km,
                "length_m": length_m,
                "lhs_width_m": _num(row[4]) if len(row) > 4 else None,
                "rhs_width_m": _num(row[5]) if len(row) > 5 else None,
                "median_superstructure": _text(row[6]) if len(row) > 6 else None,
                "span_arrangement": _text(row[7]) if len(row) > 7 else None,
                "vertical_clearance_m": _num(row[8]) if len(row) > 8 else None,
                "skew_angle": _text(row[9]) if len(row) > 9 else None,
                "remarks": _text(row[10]) if len(row) > 10 else None,
            }
        )
    return items


def _parse_interchanges(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        first = str(row[0] or "").strip().lower()
        if first.startswith("note:") or first.startswith("details of ramps"):
            break
        if not _row_has_serial(row):
            continue
        chainage = _num(row[1]) if len(row) > 1 else None
        name = _text(row[2]) if len(row) > 2 else None
        if chainage is None or not name:
            continue
        items.append(
            {
                "chainage_km": round(chainage, 3),
                "name": name,
                "span_arrangement": _text(row[3]) if len(row) > 3 else None,
                "total_width_m": _text(row[4]) if len(row) > 4 else None,
                "tcs": _text(row[5]) if len(row) > 5 else None,
                "remarks": _text(row[6]) if len(row) > 6 else None,
            }
        )
    return items


def _parse_interchange_ramps(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    current_ic: str | None = None
    started = False
    for row in rows:
        first = str(row[0] or "").strip()
        if first.startswith("Details of Ramps"):
            started = True
            continue
        if not started:
            continue
        if first and (
            "Interchange at" in first or "Trumpet at" in first or "Clover" in first
        ):
            current_ic = first
            continue
        if not _row_has_serial(row):
            continue
        from_km = _num(row[1]) if len(row) > 1 else None
        to_km = _num(row[2]) if len(row) > 2 else None
        if from_km is None or to_km is None:
            continue
        items.append(
            {
                "interchange": current_ic,
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "carriageway_width_m": _num(row[3]) if len(row) > 3 else None,
                "length_km": _num(row[4]) if len(row) > 4 else None,
                "description": _text(row[5]) if len(row) > 5 else None,
                "remarks": _text(row[6]) if len(row) > 6 else None,
            }
        )
    return items


def _parse_drains(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        from_km = _num(row[1]) if len(row) > 1 else None
        to_km = _num(row[2]) if len(row) > 2 else None
        if from_km is None or to_km is None:
            continue
        items.append(
            {
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "lhs_length_km": _num(row[3]) if len(row) > 3 else None,
                "rhs_length_km": _num(row[4]) if len(row) > 4 else None,
                "width": _text(row[5]) if len(row) > 5 else None,
                "total_length_km": _num(row[6]) if len(row) > 6 else None,
            }
        )
    return items


def _parse_culverts(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        chainage = _num(row[1]) if len(row) > 1 else None
        if chainage is None:
            continue
        span = _text(row[2]) if len(row) > 2 else None
        if not span:
            continue
        items.append(
            {
                "chainage_km": round(chainage, 3),
                "span_arrangement": span,
                "type": _text(row[3]) if len(row) > 3 else None,
                "remarks": _text(row[4]) if len(row) > 4 else None,
            }
        )
    return items


def _parse_re_walls(rows: list[tuple]) -> list[dict]:
    items: list[dict] = []
    for row in rows:
        if not _row_has_serial(row):
            continue
        from_km = _num(row[2]) if len(row) > 2 else None
        to_km = _num(row[3]) if len(row) > 3 else None
        if from_km is None or to_km is None:
            continue
        items.append(
            {
                "chainage_km": _num(row[1]) if len(row) > 1 else None,
                "from_km": round(from_km, 3),
                "to_km": round(to_km, 3),
                "lhs_length_km": _num(row[4]) if len(row) > 4 else None,
                "rhs_length_km": _num(row[5]) if len(row) > 5 else None,
                "total_length_km": _num(row[6]) if len(row) > 6 else None,
                "remarks": _text(row[7]) if len(row) > 7 else None,
            }
        )
    return items


def _build_structure_inventory(data: dict) -> list[dict]:
  """Flatten point structures for map markers and dashboard tables."""
  inventory: list[dict] = []

  for item in data.get("underpasses", []):
      inventory.append(
          {
              "type": "underpass",
              "label": item.get("remarks") or "Underpass",
              "chainage_km": item["chainage_km"],
              "details": item,
          }
      )

  for item in data.get("overpasses", []):
      inventory.append(
          {
              "type": "overpass",
              "label": item.get("remarks") or "Overpass",
              "chainage_km": item["chainage_km"],
              "details": item,
          }
      )

  for item in data.get("interchanges", []):
      if item.get("chainage_km", 0) == 0 and (item.get("name") or "").lower().startswith("as per"):
          continue
      inventory.append(
          {
              "type": "interchange",
              "label": item.get("name") or "Interchange",
              "chainage_km": item["chainage_km"],
              "details": item,
          }
      )

  for item in data.get("culverts", []):
      inventory.append(
          {
              "type": "culvert",
              "label": item.get("remarks") or "Culvert",
              "chainage_km": item["chainage_km"],
              "details": item,
          }
      )

  for item in data.get("elevated", []):
      inventory.append(
          {
              "type": "elevated",
              "label": "Elevated Section",
              "chainage_km": round((item["from_km"] + item["to_km"]) / 2, 3),
              "details": item,
          }
      )

  inventory.sort(key=lambda x: x["chainage_km"])
  return inventory


def _build_summary(data: dict) -> dict:
    tcs = data.get("tcs_stretches", [])
    main_tcs = [
        t
        for t in tcs
        if t.get("tcs")
        and t["to_km"] <= 36
        and not (t.get("remarks") or "").lower().startswith("bhita")
    ]
    centreline_km = max((t["to_km"] for t in main_tcs), default=35.0)

    interchanges = [
        i
        for i in data.get("interchanges", [])
        if not (
            i.get("chainage_km") == 0
            and (i.get("name") or "").lower().startswith("as per")
        )
    ]

    counts = {
        "underpasses": len(data.get("underpasses", [])),
        "overpasses": len(data.get("overpasses", [])),
        "interchanges": len(interchanges),
        "culverts": len(data.get("culverts", [])),
        "re_walls": len(data.get("re_walls", [])),
        "drain_sections": len(data.get("drains", [])),
        "service_road_sections": len(data.get("service_roads", [])),
        "elevated_sections": len(data.get("elevated", [])),
        "tcs_stretches": len(main_tcs),
        "interchange_ramps": len(data.get("interchange_ramps", [])),
    }
    counts["total_structures"] = (
        counts["underpasses"]
        + counts["overpasses"]
        + counts["interchanges"]
        + counts["culverts"]
    )

    return {
        "centreline_km": round(centreline_km, 3),
        "counts": counts,
    }
