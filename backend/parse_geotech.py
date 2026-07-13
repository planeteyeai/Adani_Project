"""Parse geotechnical borehole / soil-test workbooks into JSON for the frontend."""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

# Complete soil analysis workbook (locations + layer-wise tests for BH 01–35).
SOIL_ANALYSIS_SRC = Path(r"C:\Users\Kunal.Desale\Downloads\c524e929f54b4dd28931c3ad9c48af8d.xlsx")
OUT = Path(r"C:\Users\Kunal.Desale\Desktop\Adani\frontend\public\geotech_boreholes.json")


def dms_to_dd(s: str) -> list[float]:
    parts = re.findall(
        r"([0-9]+(?:\.[0-9]+)?)\D+([0-9]+(?:\.[0-9]+)?)\D+([0-9]+(?:\.[0-9]+)?)\D*([NSEW])",
        s,
    )
    out: list[float] = []
    for d, m, sec, hemi in parts:
        dd = float(d) + float(m) / 60 + float(sec) / 3600
        if hemi in ("S", "W"):
            dd = -dd
        out.append(round(dd, 6))
    return out


def clean_header(h: object) -> str:
    if h is None:
        return ""
    text = str(h).strip()
    # Excel often stores ² as a private/replacement glyph — normalise to ASCII-safe labels
    # that match the frontend BoreholeCard keys.
    text = re.sub(r"UCS\s*\(kg/cm.?\)", "UCS (kg/cm²)", text, flags=re.I)
    text = re.sub(r"SBC\s*\(T/m.?\)", "SBC (T/m²)", text, flags=re.I)
    text = text.replace("kg/cm2", "kg/cm²").replace("T/m2", "T/m²")
    return text


def clean_depth(v: object) -> str:
    if v is None:
        return ""
    text = str(v).strip()
    text = text.replace("�", "–").replace("-", "–")
    # Collapse accidental double dashes
    text = re.sub(r"–+", "–", text)
    return text


def parse_locations(ws) -> dict[int, tuple[float, float]]:
    """Sheet1: Location NN + DMS coordinate string."""
    coords: dict[int, tuple[float, float]] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        name = str(row[0]).strip()
        m = re.match(r"Location\s*0?(\d+)$", name, re.I)
        if not m or row[1] is None:
            continue
        num = int(m.group(1))
        dd = dms_to_dd(str(row[1]))
        if len(dd) == 2:
            coords[num] = (dd[0], dd[1])
    return coords


def parse_soil_layers(ws) -> tuple[dict[int, list[dict]], list[str]]:
    """Sheet2: Location headers interleaved with Depth tables."""
    by_num: dict[int, list[dict]] = {}
    headers: list[str] = []
    cur_num: int | None = None

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        c0 = str(row[0]).strip()

        loc = re.match(r"Location\s*0?(\d+)\s*:", c0, re.I)
        if loc:
            cur_num = int(loc.group(1))
            by_num.setdefault(cur_num, [])
            continue

        if c0.lower().startswith("depth"):
            headers = [clean_header(x) for x in row]
            continue

        if cur_num is None or not headers:
            continue
        if not re.match(r"^[0-9]", c0):
            continue

        rec: dict = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            if h == "Depth (m)":
                rec[h] = clean_depth(v)
            elif isinstance(v, str):
                rec[h] = v.strip()
            else:
                rec[h] = v
        by_num[cur_num].append(rec)

    return by_num, [h for h in headers if h]


def main() -> None:
    if not SOIL_ANALYSIS_SRC.exists():
        print(f"Missing workbook: {SOIL_ANALYSIS_SRC}", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(SOIL_ANALYSIS_SRC, data_only=True)
    # Prefer named sheets; fall back to first two sheets (Sheet1 / Sheet2).
    loc_ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    soil_ws = wb["Sheet2"] if "Sheet2" in wb.sheetnames else wb[wb.sheetnames[1]]

    coords = parse_locations(loc_ws)
    soil_by_num, headers = parse_soil_layers(soil_ws)

    all_nums = sorted(set(coords) | set(soil_by_num))
    boreholes = []
    for num in all_nums:
        lat = lon = None
        if num in coords:
            lat, lon = coords[num]
        boreholes.append(
            {
                "id": f"BH-{num:02d}",
                "name": f"Location {num:02d}",
                "lat": lat,
                "lon": lon,
                "layers": soil_by_num.get(num, []),
            }
        )

    payload = {
        "title": "Geotechnical Investigation — Boreholes & Soil Test Summary",
        "source_file": SOIL_ANALYSIS_SRC.name,
        "count": len(boreholes),
        "detailed_count": sum(1 for b in boreholes if b["layers"]),
        "columns": headers,
        "boreholes": boreholes,
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    detailed = [b["id"] for b in boreholes if b["layers"]]
    total_layers = sum(len(b["layers"]) for b in boreholes)
    print(f"boreholes: {len(boreholes)}  with soil detail: {len(detailed)}  layers: {total_layers}")
    print("detailed IDs:", ", ".join(detailed))
    print("columns:", headers)
    print("written", OUT)

    missing = [b["id"] for b in boreholes if b["lat"] is None]
    if missing:
        print("MISSING COORDS:", ", ".join(missing), file=sys.stderr)


if __name__ == "__main__":
    main()
